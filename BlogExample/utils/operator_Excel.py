#!/usr/bin/env python
# -*- coding: utf-8 -*-
# -------------------------------------------------------------------------------
# File:         operator_Excel.py
# Author:       ls
# Date:         2020/10/12 17:56
# -------------------------------------------------------------------------------

from openpyxl import load_workbook
from openpyxl.styles import colors, Side, Border
from openpyxl.styles import Font
# from conf.data_conf import *
import os


class OPExcel:
    def __init__(self, filepath):
        self.filepath = filepath
        # self.ws = None
        self.wb = None
        if not os.path.exists(filepath) or not (".xlsx" in filepath):
            print("输入的文件 %s 不存在或者文件类型不是.xlsx格式 " % self.filepath)
        else:
            self.wb = load_workbook(self.filepath)
            # 显示最后保存时的sheet页对象
            self.ws = self.wb.active

    # 获取所有的sheet
    def get_sheetnames(self):
        return self.wb.sheetnames

    # 获得最大的行数
    def get_max_rows(self):
        return self.ws.max_row

    # 获得最大列数
    def get_max_cols(self):
        return self.ws.max_column

    # 已经知道sheetname,在设置一次self.sheet,意义在于显式切换sheet页
    def set_sheet_by_sheetname(self, sheetname):
        self.ws = self.wb[sheetname]

    # 通过下标获取对应sheet名，下标从1开始
    def set_sheet_by_index(self, index):
        if not isinstance(index, int):
            print("您设定的sheet序号 %s 不是整数，请重新设定" % index)
            return
        elif index > len(self.get_sheetnames()) or index <= 0:
            print("您设定的sheet序号 %s 不存在，请重新设定" % index)
            return
        else:
            sheet = self.get_sheetnames()[index - 1]
            self.ws = self.wb[sheet]
            return self.ws

    # 默认获取第一个sheet内容，可接受指定第几个sheet
    def getExcelContent(self):

        # print(self.ws[1][1].value)
        # print(self.wb.active)
        res = []
        for row in self.ws.iter_rows():
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            res.append(row_data)
        return res

    # 根据指定的行列写入单元格内容
    def writeContent(self, row, col, data):
        self.ws[row][col].value = data
        if data == "fail":
            self.ws[row][col].font = Font(color=colors.RED)
        else:
            self.ws[row][col].font = Font(color=colors.BLACK)
        self.wb.save(self.filepath)

    # 设置单元格边框
    def set_cell_style(self):
        bd=Side(style='thin',color='000000')
        for row in self.ws.rows:
            for cell in row:
                cell.border=Border(left=bd,top=bd,right=bd,bottom=bd)
        self.wb.save(self.filepath)

    # 写入某一列的数据
    def write_col_data(self,col_no,data):
        for i in range(1,self.get_max_rows()):
            # self.ws[i+1][col_no-1].value=data
            self.writeContent(i+1,col_no-1,data)

    # 获取某一行的行对象
    def get_a_line(self,row_no):
        if not isinstance(row_no,int):
            print("输入的行号%s只能是整数" %row_no)
            return
        if not 0<row_no<self.ws.max_row+1:
            print("输入的行号%s超过行数范围" % row_no)
            return

        data=[]
        for row in self.ws.iter_rows():
            data.append(row)

        return data[row_no-1]

    # 获取某行的数据
    def get_line_data(self,row_no):
        # # 获取某一行的数据
        # return self.getExcelContent()[row_no-1]

        value=[]
        for cell in self.get_a_line(row_no):
            value.append(cell.value)

        return value

    # 获取指定列的值
    def get_col_data(self,col_no):
        if not isinstance(col_no, int):
            print("输入的列号%s只能是整数" % col_no)
            return
        if not 0 < col_no < self.ws.max_column + 1:
            print("输入的列号%s超过行数范围" % col_no)
            return

        data=[]
        for col in self.ws.iter_cols():
            tmp=[]
            for cell in col:
                tmp.append(cell.value)
            data.append(tmp)
        return data[col_no-1]

    # 获取每个单元格的值
    def get_cell_value(self,row_no,col_no):
        if not isinstance(row_no,int) or (not isinstance(col_no,int)):
            print("必须输入整数类型的行号或者列号")
            return

        if not 0<row_no<self.get_max_rows():
            print("输入的行号超过范围")
            return
        if not 0<col_no<self.get_max_cols():
            print("输入的列号超过范围")
            return

        return self.ws[row_no][col_no-1].value


if __name__ == "__main__":
    pass
    